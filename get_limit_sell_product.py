import os
import shutil
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from common_utils import jy_ap_col, self_operation


def get_limit_sell_product(r_dir, xsg_p, xsg_o):
    file_type = ".xlsx"
    key_word = "股票交易安排"
    data_dir = os.path.join(r_dir, xsg_p)
    pb_file_name = "pb_list.xlsx"
    file_list = get_file_list(data_dir, key_word, file_type)
    week_dir = os.path.join(r_dir, "output")
    xsg_dir = os.path.join(r_dir, xsg_o)
    week_key = "新股卖出分配周报"
    daily_key = "新股卖出分配日报"

    if not file_list:
        return False
    # 每次检查一个以往的交易安排来确定限售股
    elif len(file_list) != 1:
        print(print_info("E"), end=" ")
        print("Only one limit sell plan can be run!")
        return False

    df_zero = pd.DataFrame(columns=jy_ap_col)

    for file in file_list:
        file_path = os.path.join(data_dir, file)
        df = get_single_df(file_path, jy_ap_col)
        if type(df) == bool and df is False:
            return False
        df_zero = pd.concat([df_zero, df]).drop_duplicates()

    check_list = df_zero["账户"].tolist()
    pb_list = get_pb_list(root_dir, pb_file_name)

    xsg_list = list()
    need_product_list = list()
    # 查看新股解禁中没有进PB也不是客户自行交易的产品
    print(print_info(), end=" ")
    print("Check list:")
    for check_item in check_list:
        if check_item not in pb_list:
            # 扣除客户自行交易
            if check_item not in self_operation:
                print(check_item)
                # 每一个元素 [账户, 账号, 方案]
                xsg_item = [
                    check_item,
                    df_zero.loc[df_zero[jy_ap_col[0]] == check_item].iloc[0][jy_ap_col[1]],
                    df_zero.loc[df_zero[jy_ap_col[0]] == check_item].iloc[0][jy_ap_col[2]]
                ]
                need_product_list.append(check_item)
                xsg_list.append(xsg_item)

    # 获取最新的周报数据
    week_list = get_file_list(week_dir, week_key, file_type, False)
    if len(week_list) == 0:
        print(print_info("E"), end=" ")
        print("Can not find the week report!")
        return False

    # 复制一个周报数据
    print(print_info(), end=" ")
    print("Get week report: {}".format(week_list[-1]))
    op_date = time.strftime('%Y%m%d', time.localtime(time.time()))
    daily_name = daily_key + op_date + file_type
    shutil.copy(
        os.path.join(week_dir, week_list[-1]),
        os.path.join(xsg_dir, daily_name)
    )

    # 获取周报数据
    wb = load_workbook(filename=os.path.join(xsg_dir, daily_name))
    print(wb.sheetnames)
    sheet = wb[wb.sheetnames[0]]
    max_row = sheet.max_row

    # 获取周报中的产品列表
    df = pd.read_excel(os.path.join(xsg_dir, daily_name))
    product_list = df[jy_ap_col[0]].tolist()

    count_index = 1
    for item in xsg_list:
        if item[0] in product_list:
            # 表头占 1 行，excel 从 1 开始计数
            # print(item[0], product_list.index(item[0]) + 2)
            tmp_index = product_list.index(item[0]) + 2
            cell = sheet.cell(tmp_index, 1)
            cell.font = Font(color="008000")
            cell.fill = PatternFill(fill_type='solid', fgColor="00FA9A")
        else:
            tmp_index = max_row + count_index
            cell = sheet.cell(tmp_index, 1)
            cell.value = item[0]
            cell.font = Font(color="008000")
            cell.fill = PatternFill(fill_type='solid', fgColor="00FA9A")
            sheet.cell(tmp_index, 2).value = item[1]
            sheet.cell(tmp_index, 3).value = item[2]
            count_index += 1

    wb.save(os.path.join(xsg_dir, daily_name))
    wb.close()

    return xsg_list


def get_time(date=False, utc=False, msl=3):
    if date:
        time_fmt = "%Y-%m-%d %H:%M:%S.%f"
    else:
        time_fmt = "%H:%M:%S.%f"

    if utc:
        return datetime.utcnow().strftime(time_fmt)[:(msl-6)]
    else:
        return datetime.now().strftime(time_fmt)[:(msl-6)]


def print_info(status="I"):
    return "\033[0;33;1m[{} {}]\033[0m".format(status, get_time())


def get_file_list(data_dir, key_w, f_type=".xlsx", isprint=True):
    file_list = list()
    try:
        all_file_list = os.listdir(data_dir)
        for file in all_file_list:
            if key_w in file and "$" not in file and f_type.split(".")[-1] == file.split(".")[-1]:
                file_list.append(file)
    except FileNotFoundError as e:
        print(print_info("E"), end=" ")
        print("{}: {} can not found!".format(e, data_dir))
        return False
    except:
        print(print_info("E"))
        print("Unknown error happened!")
        return False
    finally:
        if isprint:
            print(print_info(), end=" ")
            print("Operating file set is {}".format(file_list))
        return file_list


def get_single_df(file_n, col_lst):
    try:
        df = pd.read_excel(file_n, header=1, converters={0: str})[col_lst]
        df = df[df[col_lst[2]].notna()]
        df[col_lst[-1]].fillna("PB", inplace=True)
        print(print_info(), end=" ")
        print("Loaded the {}".format(file_n))
        return df
    except:
        print(print_info("E"), " ")
        print("Can not loading the {}".format(file_n))
        return False


def get_pb_list(r_path, pb_file_name):
    pb_file_path = os.path.join(r_path, pb_file_name)
    return pd.read_excel(pb_file_path, header=None, index_col=None)[0].tolist()


if __name__ == '__main__':
    root_dir = os.path.abspath(".")
    xsg_path = "xsg_plan"
    xsg_output = "xsg_output"
    TF = get_limit_sell_product(root_dir, xsg_path, xsg_output)
    if TF:
        print(print_info("S"), end=" ")
        print("Success!")
    else:
        print(print_info("E"), end=" ")
        print("Error!")
